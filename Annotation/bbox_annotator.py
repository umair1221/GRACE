#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
BBox Annotator — Task-wise image labeling GUI (with Excel/CSV metadata + Zoom/Pan)
----------------------------------------------------------------------------------
- Load a dataset ROOT directory containing Task-* folders.
- Optionally load a metadata file (CSV or XLSX) with columns:
    Id | Query | Instructions
  where:
    * Id          : integer task number (e.g., 1 for Task-1)
    * Query       : task prompt to store for that task
    * Instructions: multi-step text (e.g., "1) Do X\n2) Do Y"), each step maps
                    to an image inside the task folder. The mapping uses either:
                    - A number found in the image filename (e.g., step_2.png → step=2), or
                    - The sorted image order (1..N) if no number is found.
- Draw bounding boxes (left-click & drag), select by clicking inside, Delete to remove.
- Per-image "Instruction" box (pre-filled from metadata; editable).
- Zoom (mouse wheel or Ctrl-'+'/'-') and pan (right-button drag).
- Save per-task JSON -> <Task-X>/annotations.json
- Export all tasks -> <ROOT>/all_annotations.json

Dependencies:
  - Python 3.8+
  - pillow (PIL)
  - openpyxl (only if you want to load .xlsx directly; otherwise export metadata to .csv)
"""

import os
import re
import json
import sys
from datetime import datetime
from typing import List, Dict, Optional, Tuple

try:
    import tkinter as tk
    from tkinter import ttk, filedialog, messagebox
except Exception as e:
    print("Tkinter is required but not available:", e)
    sys.exit(1)

try:
    from PIL import Image, ImageTk
except ImportError:
    print("Pillow is required. Please run: pip install pillow")
    sys.exit(1)

# Optional XLSX support
HAVE_OPENPYXL = False
try:
    import openpyxl  # type: ignore
    HAVE_OPENPYXL = True
except Exception:
    HAVE_OPENPYXL = False


IMAGE_EXTS = {".jpg", ".jpeg", ".png", ".bmp", ".tif", ".tiff"}


def now_iso() -> str:
    return datetime.utcnow().replace(microsecond=0).isoformat() + "Z"


def list_subdirs(path: str) -> List[str]:
    try:
        return sorted([d for d in os.listdir(path) if os.path.isdir(os.path.join(path, d))])
    except Exception:
        return []


def list_images(path: str) -> List[str]:
    try:
        files = sorted(
            [
                f for f in os.listdir(path)
                if os.path.isfile(os.path.join(path, f)) and os.path.splitext(f)[1].lower() in IMAGE_EXTS
            ],
            key=naturalsort_key
        )
        return files
    except Exception:
        return []


def naturalsort_key(s: str):
    return [int(t) if t.isdigit() else t.lower() for t in re.split(r'(\d+)', s)]


def extract_first_int(s: str) -> Optional[int]:
    m = re.search(r'(\d+)', s)
    return int(m.group(1)) if m else None


def parse_instructions(text: str) -> List[str]:
    """Split a multi-step instructions string into steps."""
    if not text:
        return []
    t = str(text).strip().replace('\r\n', '\n').replace('\r', '\n')
    lines = [ln.strip() for ln in t.split('\n') if ln.strip()]
    if len(lines) > 1:
        return lines
    chunks = re.split(r'(?:^|\s)(?:\d+[\)\.])\s*', t)
    steps = [c.strip() for c in chunks if c.strip()]
    if len(steps) > 1:
        return steps
    return [t]


# ----------------------------- Metadata ---------------------------------

class Metadata:
    def __init__(self):
        self.by_task: Dict[str, Dict] = {}

    def load_csv(self, path: str):
        import csv
        with open(path, "r", encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            self._consume_rows(list(reader))

    def load_xlsx(self, path: str):
        if not HAVE_OPENPYXL:
            raise RuntimeError("openpyxl is not installed. Run: pip install openpyxl")
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        headers = [str(c.value).strip() if c.value is not None else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
        rows = []
        for row in ws.iter_rows(min_row=2):
            data = {}
            for h, cell in zip(headers, row):
                data[h] = "" if cell.value is None else str(cell.value)
            rows.append(data)
        self._consume_rows(rows)

    def _consume_rows(self, rows: List[Dict[str, str]]):
        self.by_task.clear()
        def get(d, *keys):
            for k in keys:
                if k in d:
                    return d[k]
                for kk in d.keys():
                    if kk.strip().lower() == k.strip().lower():
                        return d[kk]
            return ""
        for r in rows:
            id_str = get(r, "Id", "Task", "TaskId", "Task ID")
            query = get(r, "Query", "Task", "Tasks", "Prompt") or ""
            instructions = get(r, "Instructions", "Steps", "Step", "Guide") or ""
            if not id_str:
                continue
            try:
                task_num = int(re.findall(r'\d+', id_str)[0])
            except Exception:
                task_num = extract_first_int(id_str) or None
            if not task_num:
                continue
            task_name = f"Task-{task_num}"
            steps = parse_instructions(instructions)
            self.by_task[task_name] = {"query": query.strip(), "steps": {i+1: s for i, s in enumerate(steps)}}

    def get_query(self, task_name: str) -> str:
        return self.by_task.get(task_name, {}).get("query", "")

    def get_instruction_for_image(self, task_name: str, filename: str, ordinal_index1: int) -> Tuple[int, str]:
        steps_map = self.by_task.get(task_name, {}).get("steps", {})
        k = extract_first_int(filename) or ordinal_index1
        return k, steps_map.get(k, "")


# ----------------------------- Annotation Store --------------------------

class AnnotationStore:
    def __init__(self, dataset_root: str):
        self.dataset_root = dataset_root
        self.data: Dict[str, Dict] = {}

    def _json_path(self, task_path: str) -> str:
        return os.path.join(task_path, "annotations.json")

    def load_task(self, task_path: str, task_name: str):
        jp = self._json_path(task_path)
        if os.path.exists(jp):
            try:
                with open(jp, "r", encoding="utf-8") as f:
                    data = json.load(f)
                data.setdefault("task_name", task_name)
                data.setdefault("task_query", "")
                data.setdefault("dataset_root", self.dataset_root)
                data["updated_at"] = now_iso()
                data.setdefault("images", [])
                for img in data["images"]:
                    img.setdefault("step_index", 0)
                    img.setdefault("instruction", "")
                self.data[task_path] = data
            except Exception:
                self._fresh(task_path, task_name)
        else:
            self._fresh(task_path, task_name)

    def _fresh(self, task_path: str, task_name: str):
        self.data[task_path] = {
            "task_name": task_name,
            "task_query": "",
            "dataset_root": self.dataset_root,
            "updated_at": now_iso(),
            "images": []
        }

    def ensure_image_entry(self, task_path: str, file_name: str, w: int, h: int) -> Dict:
        task_data = self.data.get(task_path)
        if task_data is None:
            raise RuntimeError("Task not loaded")
        for img in task_data["images"]:
            if img.get("file_name") == file_name:
                img["width"], img["height"] = w, h
                img.setdefault("step_index", 0)
                img.setdefault("instruction", "")
                img.setdefault("bboxes", [])
                return img
        entry = {"file_name": file_name, "width": w, "height": h, "step_index": 0, "instruction": "", "bboxes": []}
        task_data["images"].append(entry)
        return entry

    def get_image_entry(self, task_path: str, file_name: str) -> Optional[Dict]:
        task_data = self.data.get(task_path)
        if not task_data:
            return None
        for img in task_data["images"]:
            if img.get("file_name") == file_name:
                return img
        return None

    def set_task_query(self, task_path: str, query_text: str):
        td = self.data.get(task_path)
        if td is not None:
            td["task_query"] = query_text

    def save_task(self, task_path: str):
        td = self.data.get(task_path)
        if td is None:
            return
        td["updated_at"] = now_iso()
        jp = self._json_path(task_path)
        with open(jp, "w", encoding="utf-8") as f:
            json.dump(td, f, indent=2, ensure_ascii=False)

    def export_all(self, out_path: str, task_paths: List[str]):
        payload = {"exported_at": now_iso(), "dataset_root": self.dataset_root, "tasks": []}
        for tp in task_paths:
            if tp not in self.data:
                self.load_task(tp, os.path.basename(tp))
            payload["tasks"].append(self.data[tp])
        with open(out_path, "w", encoding="utf-8") as f:
            json.dump(payload, f, indent=2, ensure_ascii=False)


# ----------------------------- Image Canvas ------------------------------

class ImageCanvas(tk.Canvas):
    def __init__(self, master, width=1040, height=720, **kwargs):
        super().__init__(master, width=width, height=height, bg="#202225", highlightthickness=0, **kwargs)
        self.img_tk: Optional[ImageTk.PhotoImage] = None
        self.img: Optional[Image.Image] = None
        self.img_path: Optional[str] = None
        self.img_w: int = 0
        self.img_h: int = 0

        self.pad = 8
        self.fit_scale = 1.0
        self.zoom = 1.0
        self.offset_x = 0
        self.offset_y = 0

        self.image_item_id = None
        self.rects: List[Dict] = []
        self._next_rect_id = 1
        self._selected_idx: Optional[int] = None

        self._drawing = False
        self._start_x = 0
        self._start_y = 0
        self._temp_rect_id = None

        self._panning = False
        self._pan_start = (0, 0)
        self._orig_offset = (0, 0)

        self.bind("<Button-1>", self.on_mouse_down_left)
        self.bind("<B1-Motion>", self.on_mouse_drag_left)
        self.bind("<ButtonRelease-1>", self.on_mouse_up_left)

        self.bind("<Button-3>", self.on_pan_start)
        self.bind("<B3-Motion>", self.on_pan_move)
        self.bind("<ButtonRelease-3>", self.on_pan_end)

        self.bind("<MouseWheel>", self.on_wheel)             # Windows
        self.bind("<Control-MouseWheel>", self.on_wheel_ctrl)
        self.bind("<Button-4>", lambda e: self._zoom(1.1, e.x, e.y))  # mac
        self.bind("<Button-5>", lambda e: self._zoom(0.9, e.x, e.y))  # mac

    def current_scale(self) -> float:
        return self.fit_scale * self.zoom

    def _img_to_canvas(self, x: float, y: float) -> Tuple[float, float]:
        s = self.current_scale()
        return self.offset_x + x * s, self.offset_y + y * s

    def _canvas_to_img(self, cx: float, cy: float) -> Tuple[float, float]:
        s = self.current_scale()
        return (cx - self.offset_x) / s, (cy - self.offset_y) / s

    def set_image(self, path: str):
        self.delete("all")
        self.img_path = path
        self.img = Image.open(path).convert("RGB")
        self.img_w, self.img_h = self.img.size

        cw = int(self["width"]); ch = int(self["height"])
        if cw <= 0 or ch <= 0:
            cw, ch = 1040, 720
            self.config(width=cw, height=ch)
        scale_w = (cw - 2*self.pad) / self.img_w
        scale_h = (ch - 2*self.pad) / self.img_h
        self.fit_scale = min(scale_w, scale_h, 1.0)
        self.zoom = 1.0

        disp_w = int(self.img_w * self.current_scale())
        disp_h = int(self.img_h * self.current_scale())
        self.offset_x = (cw - disp_w)//2
        self.offset_y = (ch - disp_h)//2

        self._redraw_all()
        self.rects.clear()
        self._selected_idx = None
        self._next_rect_id = 1

    def _redraw_all(self):
        self.delete("all")
        if self.img is None:
            return
        disp_w = max(1, int(self.img_w * self.current_scale()))
        disp_h = max(1, int(self.img_h * self.current_scale()))
        disp_img = self.img.resize((disp_w, disp_h), Image.BILINEAR) if self.current_scale() != 1.0 else self.img
        self.img_tk = ImageTk.PhotoImage(disp_img)
        self.image_item_id = self.create_image(self.offset_x, self.offset_y, image=self.img_tk, anchor="nw", tags="image")
        for i, r in enumerate(self.rects):
            x, y, w, h = r["bbox"]
            selected = (self._selected_idx == i)
            r["canvas_id"] = self._draw_rect_on_canvas(x, y, w, h, selected)

    def _draw_rect_on_canvas(self, x, y, w, h, selected):
        x0, y0 = self._img_to_canvas(x, y)
        x1, y1 = self._img_to_canvas(x + w, y + h)
        outline = "#00FF88" if selected else "#00B3FF"
        width = 3 if selected else 2
        return self.create_rectangle(x0, y0, x1, y1, outline=outline, width=width, tags="rect")

    def load_rects(self, rect_list: List[Dict]):
        self.clear_rects()
        for r in rect_list:
            rid = r.get("id", self._next_rect_id)
            self._next_rect_id = max(self._next_rect_id, rid + 1)
            x, y, w, h = r.get("bbox", [0, 0, 0, 0])
            cid = self._draw_rect_on_canvas(x, y, w, h, selected=False)
            self.rects.append({"id": rid, "label": r.get("label", "object"), "bbox": [x, y, w, h], "canvas_id": cid})
        self._selected_idx = None

    def get_rects(self) -> List[Dict]:
        return [{"id": r["id"], "label": r.get("label", "object"), "bbox": list(r["bbox"])} for r in self.rects]

    def clear_rects(self):
        for r in self.rects:
            try:
                self.delete(r.get("canvas_id", None))
            except Exception:
                pass
        self.rects.clear()
        self._selected_idx = None

    def _find_rect_at_canvas_point(self, cx, cy):
        for idx, r in enumerate(self.rects):
            x, y, w, h = r["bbox"]
            rx0, ry0 = self._img_to_canvas(x, y)
            rx1, ry1 = self._img_to_canvas(x + w, y + h)
            if rx0 <= cx <= rx1 and ry0 <= cy <= ry1:
                return idx
        return None

    def select_rect(self, idx: Optional[int]):
        if self._selected_idx is not None and 0 <= self._selected_idx < len(self.rects):
            r = self.rects[self._selected_idx]
            try:
                self.delete(r["canvas_id"])
            except Exception:
                pass
            x, y, w, h = r["bbox"]
            r["canvas_id"] = self._draw_rect_on_canvas(x, y, w, h, selected=False)

        self._selected_idx = idx
        if idx is not None and 0 <= idx < len(self.rects):
            r = self.rects[idx]
            try:
                self.delete(r["canvas_id"])
            except Exception:
                pass
            x, y, w, h = r["bbox"]
            r["canvas_id"] = self._draw_rect_on_canvas(x, y, w, h, selected=True)

    def delete_selected(self):
        if self._selected_idx is None:
            return
        idx = self._selected_idx
        try:
            self.delete(self.rects[idx]["canvas_id"])
        except Exception:
            pass
        del self.rects[idx]
        self._selected_idx = None

    def on_mouse_down_left(self, event):
        if self.img is None:
            return
        idx = self._find_rect_at_canvas_point(event.x, event.y)
        if idx is not None:
            self.select_rect(idx)
            self._drawing = False
            return
        self._drawing = True
        self._start_x, self._start_y = event.x, event.y
        if self._temp_rect_id is not None:
            try:
                self.delete(self._temp_rect_id)
            except Exception:
                pass
            self._temp_rect_id = None

    def on_mouse_drag_left(self, event):
        if not self._drawing:
            return
        if self._temp_rect_id is not None:
            try:
                self.delete(self._temp_rect_id)
            except Exception:
                pass
        self._temp_rect_id = self.create_rectangle(
            self._start_x, self._start_y, event.x, event.y,
            outline="#FFD500", width=2, dash=(4, 2), tags="temp"
        )

    def on_mouse_up_left(self, event):
        if not self._drawing:
            return
        self._drawing = False
        if self._temp_rect_id is not None:
            try:
                self.delete(self._temp_rect_id)
            except Exception:
                pass
            self._temp_rect_id = None

        x0, y0 = self._start_x, self._start_y
        x1, y1 = event.x, event.y
        left, right = sorted([x0, x1])
        top, bottom = sorted([y0, y1])

        disp_x0, disp_y0 = self._img_to_canvas(0, 0)
        disp_x1, disp_y1 = self._img_to_canvas(self.img_w, self.img_h)
        left = max(left, disp_x0); top = max(top, disp_y0)
        right = min(right, disp_x1); bottom = min(bottom, disp_y1)

        if right - left < 4 or bottom - top < 4:
            return

        ix0, iy0 = self._canvas_to_img(left, top)
        ix1, iy1 = self._canvas_to_img(right, bottom)
        x, y = max(0.0, ix0), max(0.0, iy0)
        w, h = max(1.0, ix1 - ix0), max(1.0, iy1 - iy0)

        label = getattr(self.master, "get_current_label", lambda: "object")()
        cid = self._draw_rect_on_canvas(x, y, w, h, selected=False)
        self.rects.append({"id": self._next_rect_id, "label": label, "bbox": [float(x), float(y), float(w), float(h)], "canvas_id": cid})
        self._next_rect_id += 1
        self.select_rect(len(self.rects) - 1)

    def on_pan_start(self, event):
        if self.img is None:
            return
        self._panning = True
        self._pan_start = (event.x, event.y)
        self._orig_offset = (self.offset_x, self.offset_y)

    def on_pan_move(self, event):
        if not self._panning:
            return
        dx = event.x - self._pan_start[0]
        dy = event.y - self._pan_start[1]
        self.offset_x = self._orig_offset[0] + dx
        self.offset_y = self._orig_offset[1] + dy
        self._redraw_all()

    def on_pan_end(self, event):
        self._panning = False

    def on_wheel(self, event):
        factor = 1.1 if event.delta > 0 else 0.9
        self._zoom(factor, event.x, event.y)

    def on_wheel_ctrl(self, event):
        factor = 1.2 if event.delta > 0 else 1/1.2
        self._zoom(factor, event.x, event.y)

    def _zoom(self, factor: float, cx: float, cy: float):
        if self.img is None:
            return
        new_zoom = max(0.1, min(8.0, self.zoom * factor))
        factor = new_zoom / self.zoom
        if abs(factor - 1.0) < 1e-6:
            return
        img_x, img_y = self._canvas_to_img(cx, cy)
        self.zoom = new_zoom
        new_cx, new_cy = self._img_to_canvas(img_x, img_y)
        self.offset_x += (cx - new_cx)
        self.offset_y += (cy - new_cy)
        self._redraw_all()

    def zoom_reset(self):
        if self.img is None:
            return
        self.zoom = 1.0
        cw, ch = int(self["width"]), int(self["height"])
        disp_w = int(self.img_w * self.current_scale())
        disp_h = int(self.img_h * self.current_scale())
        self.offset_x = (cw - disp_w)//2
        self.offset_y = (ch - disp_h)//2
        self._redraw_all()


# ----------------------------- App --------------------------------------

class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("BBox Annotator — Task-wise GUI (with metadata)")
        self.geometry("1440x860")
        self.minsize(1180, 760)

        # State
        self.dataset_root: Optional[str] = None
        self.task_dirs: List[str] = []
        self.current_task_idx: Optional[int] = None
        self.image_files: List[str] = []
        self.current_image_idx: Optional[int] = None

        self.store: Optional[AnnotationStore] = None
        self.meta = Metadata()

        # UI vars
        self.task_query_var = tk.StringVar(value="")
        self.label_var = tk.StringVar(value="object")
        self.status_var = tk.StringVar(value="Ready")

        self.instruction_text_widget: Optional[tk.Text] = None

        # Build layout FIRST so self.canvas exists before menus reference it
        self._build_layout()
        self._build_menu()
        self._bind_shortcuts()

    # Safe menu callbacks (avoid referencing self.canvas before it exists)
    def view_zoom_in(self):
        if hasattr(self, 'canvas') and self.canvas is not None:
            self.canvas._zoom(1.2, *self._canvas_center())

    def view_zoom_out(self):
        if hasattr(self, 'canvas') and self.canvas is not None:
            self.canvas._zoom(1/1.2, *self._canvas_center())

    def view_zoom_reset(self):
        if hasattr(self, 'canvas') and self.canvas is not None:
            self.canvas.zoom_reset()

    def _build_menu(self):
        menubar = tk.Menu(self)

        filemenu = tk.Menu(menubar, tearoff=0)
        filemenu.add_command(label="Load Dataset Root...", command=self.on_load_dataset)
        filemenu.add_command(label="Load Metadata (CSV/XLSX)...", command=self.on_load_metadata)
        filemenu.add_separator()
        filemenu.add_command(label="Export all_annotations.json", command=self.on_export_all, accelerator="Ctrl+E")
        filemenu.add_separator()
        filemenu.add_command(label="Quit", command=self.on_quit)
        menubar.add_cascade(label="File", menu=filemenu)

        viewmenu = tk.Menu(menubar, tearoff=0)
        viewmenu.add_command(label="Zoom In", command=self.view_zoom_in)
        viewmenu.add_command(label="Zoom Out", command=self.view_zoom_out)
        viewmenu.add_command(label="Reset Zoom", command=self.view_zoom_reset)
        menubar.add_cascade(label="View", menu=viewmenu)

        self.config(menu=menubar)

    def _canvas_center(self) -> Tuple[int, int]:
        return (int(self.canvas.winfo_width()/2), int(self.canvas.winfo_height()/2))

    def _build_layout(self):
        root = ttk.Frame(self)
        root.pack(fill="both", expand=True)

        # Left: Task list
        left = ttk.Frame(root, padding=(8, 8))
        left.pack(side="left", fill="y")

        ttk.Label(left, text="Tasks", font=("Segoe UI", 10, "bold")).pack(anchor="w")
        self.task_list = tk.Listbox(left, width=26, height=28, exportselection=False)
        self.task_list.pack(fill="y", expand=False)
        self.task_list.bind("<<ListboxSelect>>", self.on_select_task)

        # Middle: Image list
        mid = ttk.Frame(root, padding=(8, 8))
        mid.pack(side="left", fill="y")
        ttk.Label(mid, text="Images", font=("Segoe UI", 10, "bold")).pack(anchor="w")
        self.image_list = tk.Listbox(mid, width=36, height=28, exportselection=False)
        self.image_list.pack(fill="y", expand=False)
        self.image_list.bind("<<ListboxSelect>>", self.on_select_image)

        # Right column
        right = ttk.Frame(root, padding=(8, 8))
        right.pack(side="left", fill="both", expand=True)

        # Canvas
        self.canvas = ImageCanvas(right, width=1040, height=720)
        self.canvas.pack(fill="both", expand=True)

        # Controls row
        controls = ttk.Frame(right, padding=(0, 4))
        controls.pack(fill="x")
        ttk.Label(controls, text="Current Label:").pack(side="left")
        self.label_entry = ttk.Entry(controls, textvariable=self.label_var, width=18)
        self.label_entry.pack(side="left", padx=(4, 12))

        ttk.Button(controls, text="Delete Selected (Del)", command=self.on_delete_selected).pack(side="left", padx=4)
        ttk.Button(controls, text="Prev (P)", command=self.prev_image).pack(side="left", padx=4)
        ttk.Button(controls, text="Next (N)", command=self.next_image).pack(side="left", padx=4)
        ttk.Button(controls, text="Save (S)", command=self.save_current_task).pack(side="left", padx=4)

        # Task query row
        qf = ttk.Frame(right, padding=(0, 6))
        qf.pack(fill="x")
        ttk.Label(qf, text="Task Query:").pack(side="left")
        self.task_query_entry = ttk.Entry(qf, textvariable=self.task_query_var)
        self.task_query_entry.pack(side="left", fill="x", expand=True, padx=8)
        ttk.Button(qf, text="Save Query", command=self.on_save_query).pack(side="left")

        # Instruction box for current image
        ib = ttk.Frame(right, padding=(0, 4))
        ib.pack(fill="x")
        ttk.Label(ib, text="Instruction for this image:").pack(anchor="w")
        self.instruction_text_widget = tk.Text(ib, height=3, wrap="word")
        self.instruction_text_widget.pack(fill="x", expand=False, pady=(2, 6))

        # Status bar
        status = ttk.Frame(self, relief="sunken", padding=(6, 2))
        status.pack(side="bottom", fill="x")
        self.status_label = ttk.Label(status, textvariable=self.status_var, anchor="w")
        self.status_label.pack(side="left", fill="x", expand=True)

        # Expose label getter to canvas
        self.canvas.master.get_current_label = self.get_current_label

    def _bind_shortcuts(self):
        self.bind("<Escape>", lambda e: self.focus_set())
        self.bind("<Delete>", lambda e: self.on_delete_selected())
        self.bind("<BackSpace>", lambda e: self.on_delete_selected())
        self.bind("<s>", lambda e: self.save_current_task())
        self.bind("<S>", lambda e: self.save_current_task())
        self.bind("<n>", lambda e: self.next_image())
        self.bind("<N>", lambda e: self.next_image())
        self.bind("<p>", lambda e: self.prev_image())
        self.bind("<P>", lambda e: self.prev_image())
        self.bind("<Control-e>", lambda e: self.on_export_all())
        self.bind("<Control-plus>", lambda e: self.view_zoom_in())
        self.bind("<Control-minus>", lambda e: self.view_zoom_out())
        self.protocol("WM_DELETE_WINDOW", self.on_quit)

    # -------------------- Actions ---------------------
    def on_load_dataset(self):
        path = filedialog.askdirectory(title="Select dataset ROOT (contains Task-* folders)")
        if not path:
            return
        self.dataset_root = path
        self.store = AnnotationStore(self.dataset_root)
        self._refresh_tasks()
        self.status_var.set(f"Loaded dataset: {self.dataset_root}")

    def on_load_metadata(self):
        path = filedialog.askopenfilename(
            title="Select metadata file (CSV or XLSX)",
            filetypes=[("CSV", "*.csv"), ("Excel", "*.xlsx;*.xls")]
        )
        if not path:
            return
        try:
            if path.lower().endswith(".csv"):
                self.meta.load_csv(path)
            else:
                self.meta.load_xlsx(path)
            self.status_var.set(f"Loaded metadata: {os.path.basename(path)}")
            if self.current_task_idx is not None:
                self._apply_metadata_to_current_task()
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load metadata:\\n{e}")

    def _apply_metadata_to_current_task(self):
        if self.current_task_idx is None:
            return
        task_path = self.task_dirs[self.current_task_idx]
        task_name = os.path.basename(task_path)
        auto_query = self.meta.get_query(task_name)
        if auto_query:
            self.task_query_var.set(auto_query)
            # sync store immediately
            assert self.store is not None
            self.store.set_task_query(task_path, auto_query)
            self.store.save_task(task_path)
        self._update_instruction_box_from_meta()

    def _update_instruction_box_from_meta(self):
        if self.current_task_idx is None or self.current_image_idx is None:
            return
        task_path = self.task_dirs[self.current_task_idx]
        task_name = os.path.basename(task_path)
        filename = self.image_files[self.current_image_idx]
        ordinal = self.current_image_idx + 1
        step_index, text = self.meta.get_instruction_for_image(task_name, filename, ordinal)
        assert self.store is not None
        entry = self.store.ensure_image_entry(task_path, filename, self.canvas.img_w, self.canvas.img_h)
        if not entry.get("instruction"):
            entry["step_index"] = step_index
            entry["instruction"] = text
        self._set_instruction_text(entry.get("instruction", ""))

    def _set_instruction_text(self, s: str):
        if self.instruction_text_widget is None:
            return
        self.instruction_text_widget.delete("1.0", "end")
        if s:
            self.instruction_text_widget.insert("1.0", s)

    def _refresh_tasks(self):
        assert self.dataset_root is not None
        dirs = [os.path.join(self.dataset_root, d) for d in list_subdirs(self.dataset_root)]
        task_dirs = []
        for tp in dirs:
            imgs = list_images(tp)
            if imgs:
                task_dirs.append(tp)
        task_dirs.sort(key=lambda p: naturalsort_key(os.path.basename(p)))
        self.task_dirs = task_dirs

        self.task_list.delete(0, "end")
        for tp in self.task_dirs:
            self.task_list.insert("end", os.path.basename(tp))

        self.image_list.delete(0, "end")
        self.canvas.delete("all")
        self.task_query_var.set("")
        self.current_task_idx = None
        self.current_image_idx = None

    def on_select_task(self, event=None):
        sel = self.task_list.curselection()
        if not sel:
            return
        idx = int(sel[0])
        self.current_task_idx = idx
        task_path = self.task_dirs[idx]
        task_name = os.path.basename(task_path)

        assert self.store is not None
        self.store.load_task(task_path, task_name)

        data = self.store.data.get(task_path, {})
        q_meta = self.meta.get_query(task_name)
        # Prefer metadata query if available; user may still edit afterwards
        if q_meta:
            self.task_query_var.set(q_meta)
            data["task_query"] = q_meta
            self.store.data[task_path] = data
        else:
            self.task_query_var.set(data.get("task_query", ""))

        self.image_files = list_images(task_path)
        self.image_files.sort(key=naturalsort_key)
        self.image_list.delete(0, "end")
        for f in self.image_files:
            self.image_list.insert("end", f)
        self.current_image_idx = None

        if self.image_files:
            self.image_list.selection_clear(0, "end")
            self.image_list.selection_set(0)
            self.image_list.event_generate("<<ListboxSelect>>")

        self.status_var.set(f"Task selected: {task_name} ({len(self.image_files)} images)")

    def on_select_image(self, event=None):
        sel = self.image_list.curselection()
        if not sel:
            return
        idx = int(sel[0])
        self.current_image_idx = idx

        assert self.current_task_idx is not None
        task_path = self.task_dirs[self.current_task_idx]
        img_file = self.image_files[idx]
        full_path = os.path.join(task_path, img_file)

        self.canvas.set_image(full_path)
        # Extra safety: make sure no rectangles carry over
        self.canvas.clear_rects()

        assert self.store is not None
        entry = self.store.get_image_entry(task_path, img_file)
        if entry is None:
            entry = self.store.ensure_image_entry(task_path, img_file, self.canvas.img_w, self.canvas.img_h)
            self._update_instruction_box_from_meta()
        else:
            entry["width"], entry["height"] = self.canvas.img_w, self.canvas.img_h
            self.canvas.load_rects(entry.get("bboxes", []))
            self._set_instruction_text(entry.get("instruction", ""))

        self.status_var.set(f"Image: {img_file} — {self.canvas.img_w}x{self.canvas.img_h}px")

    def on_delete_selected(self):
        self.canvas.delete_selected()

    def on_save_query(self):
        if self.current_task_idx is None:
            return
        task_path = self.task_dirs[self.current_task_idx]
        assert self.store is not None
        self.store.set_task_query(task_path, self.task_query_var.get().strip())
        self.store.save_task(task_path)
        self.status_var.set("Task query saved.")

    def get_current_label(self) -> str:
        return self.label_var.get().strip() or "object"

    def _save_current_image_entry(self):
        if self.current_task_idx is None or self.current_image_idx is None:
            return
        task_path = self.task_dirs[self.current_task_idx]
        img_file = self.image_files[self.current_image_idx]
        rects = self.canvas.get_rects()

        assert self.store is not None
        entry = self.store.ensure_image_entry(task_path, img_file, self.canvas.img_w, self.canvas.img_h)
        entry["bboxes"] = rects
        if self.instruction_text_widget is not None:
            entry["instruction"] = self.instruction_text_widget.get("1.0", "end").strip()
        if not entry.get("step_index"):
            task_name = os.path.basename(task_path)
            step_index, _ = self.meta.get_instruction_for_image(task_name, img_file, self.current_image_idx + 1)
            entry["step_index"] = step_index

    def save_current_task(self):
        if self.current_task_idx is None:
            return
        self._save_current_image_entry()
        task_path = self.task_dirs[self.current_task_idx]
        assert self.store is not None
        self.store.set_task_query(task_path, self.task_query_var.get().strip())
        self.store.save_task(task_path)
        self.status_var.set("Annotations saved for current task.")

    def next_image(self):
        if self.current_task_idx is None or not self.image_files:
            return
        if self.current_image_idx is None:
            idx = 0
        else:
            self._save_current_image_entry()
            idx = min(len(self.image_files) - 1, self.current_image_idx + 1)
        self.image_list.selection_clear(0, "end")
        self.image_list.selection_set(idx)
        self.image_list.see(idx)
        self.image_list.event_generate("<<ListboxSelect>>")

    def prev_image(self):
        if self.current_task_idx is None or not self.image_files:
            return
        if self.current_image_idx is None:
            idx = 0
        else:
            self._save_current_image_entry()
            idx = max(0, self.current_image_idx - 1)
        self.image_list.selection_clear(0, "end")
        self.image_list.selection_set(idx)
        self.image_list.see(idx)
        self.image_list.event_generate("<<ListboxSelect>>")

    def on_export_all(self):
        if not self.dataset_root:
            messagebox.showinfo("Info", "Load a dataset root first.")
            return
        self.save_current_task()
        out_path = os.path.join(self.dataset_root, "all_annotations.json")
        assert self.store is not None
        self.store.export_all(out_path, self.task_dirs)
        self.status_var.set(f"Exported: {out_path}")
        messagebox.showinfo("Export", f"Exported aggregated annotations to:\\n{out_path}")

    def on_quit(self):
        try:
            self.save_current_task()
        except Exception:
            pass
        self.destroy()


def main():
    app = App()
    app.mainloop()


if __name__ == "__main__":
    main()
